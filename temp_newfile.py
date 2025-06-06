import pandas as pd
import numpy as np
import openpyxl
import os

# 全局文件路径配置
file_path = '速卖通领星后台原始数据(3).xlsx'
output_file = '速卖通5月处理后数据.xlsx'
temp_file = '速卖通5月-temp.xlsx'
def close_wps_processes():
    try:
        # 尝试关闭WPS Office相关进程
        # 常见的WPS进程名包括：wps.exe, et.exe (表格), wpp.exe (演示), wpscloudsvr.exe (云服务)
        process_names = ['wps.exe']
        
        for p_name in process_names:
            print(f"尝试关闭进程: {p_name}...")
            # /f 强制终止进程, /im 指定镜像名称
            os.system(f"taskkill /f /im {p_name}")
            print(f"已发送关闭命令给 {p_name}")
            
    except Exception as e:
        print(f"关闭WPS进程时发生错误: {e}")

def unmerge_and_fill():
    print("\n=== 开始处理合并单元格 ===")
    print("✨ 正在读取Excel文件...")
    
    # 先用openpyxl读取以处理合并单元格
    wb = openpyxl.load_workbook(file_path)
    ws = wb['sheet1']
    
    # 获取所有合并单元格的范围并转换为列表
    merged_ranges = list(ws.merged_cells.ranges)
    
    # 对每个合并单元格进行处理
    for merged_range in merged_ranges:
        # 获取左上角的值
        top_left_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        # 取消合并
        ws.unmerge_cells(str(merged_range))
        # 填充值
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                ws.cell(row=row, column=col, value=top_left_value)
    
    # 保存临时文件
    wb.save(temp_file)
    
    # 使用pandas读取处理后的文件
    df = pd.read_excel(temp_file, sheet_name='sheet1', dtype=str)
    
    # 保存处理后的数据到输出文件（使用全局变量）
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='sheet1', index=False)
    worksheet = writer.sheets['sheet1']
    if '发货日期' in df.columns:
        df['发货日期'] = pd.to_datetime(df['发货日期'], errors='coerce').dt.strftime('%Y-%m-%d')
    for idx, col in enumerate(df.columns, 1):
        col_letter = openpyxl.utils.get_column_letter(idx)
        if col == '发货日期':
            worksheet.column_dimensions[col_letter].number_format = 'yyyy-mm-dd'
        else:
            worksheet.column_dimensions[col_letter].number_format = '@'
    
    # 删除临时文件
    if os.path.exists(temp_file):
        os.remove(temp_file)

def process_backend():
    try:
        # 读取sheet1和后台数据
        df = pd.read_excel(output_file, sheet_name='sheet1', dtype=str)
        df_backend = pd.read_excel(file_path, sheet_name='后台', dtype=str)  # 从原始文件读取后台sheet页
        
        # 处理数字格式，避免科学计数法
        def format_number(x, current_col):
            try:
                if pd.isna(x):
                        return ''
                if isinstance(x, (int, float)) or (isinstance(x, str) and any(c.isdigit() for c in str(x))):
                        num = float(str(x).strip())
                if current_col == '商品金额':
                            return '{:.2f}'.format(num)
                        # 其他数字列保持原有处理逻辑
                if isinstance(x, str) and ('e' in x.lower() or 'E' in x):
                            return '{:.2f}'.format(num)
                elif isinstance(x, (int, float)):
                            return '{:.2f}'.format(num)
                return x
            except:
                return x
        
        # 对所有列应用格式化
        for col in df.columns:
            df[col] = df[col].apply(lambda x: format_number(x, col))
        for col in df_backend.columns:
            df_backend[col] = df_backend[col].apply(lambda x: format_number(x, col))
        
        # 处理发货时间和发货日期
        if '发货时间' in df.columns:
            # 尝试以 MM/DD/YYYY 格式解析日期，如果失败则尝试其他常见格式
            df['发货日期'] = pd.to_datetime(df['发货时间'], format='%m/%d/%Y', errors='coerce')
            # 将日期格式化为 YYYY/MM/DD
            df['发货日期'] = df['发货日期'].dt.strftime('%Y/%m/%d')
        
        # 处理科学计数法的列（如果需要）
        numeric_columns = df.select_dtypes(include=['float64', 'int64']).columns
        for col in numeric_columns:
            df[col] = df[col].apply(lambda x: f'{float(x):.0f}' if pd.notnull(x) else x)
        
        # 处理平台单号
        if '平台单号' in df.columns:
            mask = df['平台单号'].str.endswith('-1', na=False)
            base_orders = df[mask]['平台单号'].str.replace('-1', '')
            df = df[~df['平台单号'].isin(base_orders) | mask]
            df['平台单号'] = df['平台单号'].str.replace('-1', '')
    
        # 定义需要的列（注意：发货日期已经在前面处理）
        column_order = [
            '发货日期', '平台单号', '系统单号', 'MSKU', 'SKU', '数量', 
            '商品金额', '商品客付运费', '发货仓库', '状态', '客服备注', 
            '运单号', '标签', '品名', '订单来源', '订单类型', 'ASIN/商品Id'
        ]
        if '商品金额' in df.columns:
            column_order.append('商品金额')
        new_columns = ['备注', '订单状态']
        
        # 筛选现有列并添加新列
        final_df = pd.DataFrame()
        
        # 添加现有列（如果存在）
        for col in column_order:
            if col in df.columns:
                final_df[col] = df[col]
            else:
                final_df[col] = ''
        
        # 添加新列（空值）
        for col in new_columns:
            final_df[col] = ''
        
        # 保存sheet1的处理结果前，将所有数据转换为文本
        def convert_scientific_notation(x):
            try:
                if pd.isna(x):
                    return ''
                x = str(x)
                # 只处理包含科学计数法的数字
                if ('e' in x.lower() or 'E' in x) and any(c.isdigit() for c in x):
                    # 确保是数字格式
                    if all(c.isdigit() or c in '.-+eE' for c in x):
                        return '{:.0f}'.format(float(x))
                return x
            except:
                return x
    
        # 对所有列应用转换
        for col in final_df.columns:
            final_df[col] = final_df[col].astype(str)
            final_df[col] = final_df[col].apply(convert_scientific_notation)
        
           
        if 'MSKU' in df.columns and '品名' in df.columns:
            # 处理MSKU的填充规则
            mask_rule1 = (df['MSKU'].isna() | (df['MSKU'] == '')) & df['品名'].notna() & (df['品名'] != '')
        df.loc[mask_rule1, 'MSKU'] = df.loc[mask_rule1, '品名']
        
        # 规则2：MSKU和品名都为空时，根据ASIN/商品Id填充特定值
        mask_rule2a = ((df['MSKU'].isna() | (df['MSKU'] == '')) & 
                      (df['品名'].isna() | (df['品名'] == '')) & 
                      (df['ASIN/商品Id'] == '1005005889378484'))
        df.loc[mask_rule2a, 'MSKU'] = 'GL-XE3000'
        
        mask_rule2b = ((df['MSKU'].isna() | (df['MSKU'] == '')) & 
                      (df['品名'].isna() | (df['品名'] == '')) & 
                      (df['ASIN/商品Id'] == '4001191579263'))
        df.loc[mask_rule2b, 'MSKU'] = '运费'
            
        # 创建最终DataFrame时按新的列顺序排序
        final_df = pd.DataFrame()
        for col in column_order:
            if col in df.columns:
                final_df[col] = df[col]
            else:
                final_df[col] = ''
    
        # 添加现有列（如果存在）
        for col in column_order:
            if col in df.columns:
                final_df[col] = df[col]
            else:
                final_df[col] = ''
        
        # 添加新列（空值）
        for col in new_columns:
            final_df[col] = ''
        
        # 保存sheet1的处理结果前检查文件是否被占用
        # output_file = '速卖通订单管理202502-处理后.xlsx'
        try:
            # 尝试打开文件检查是否可写
            with open(output_file, 'a'):
                pass
        except PermissionError:
            print(f"错误: 文件 '{output_file}' 正在被其他程序使用。请关闭该文件后重试。")
            return
            
        # 保存处理结果
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                final_df.to_excel(writer, sheet_name='sheet1', index=False)
                worksheet = writer.sheets['sheet1']
                if '发货日期' in final_df.columns:
                    final_df['发货日期'] = pd.to_datetime(final_df['发货日期'], errors='coerce').dt.strftime('%Y-%m-%d')
                for idx, col in enumerate(final_df.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    if col == '发货日期':
                        worksheet.column_dimensions[col_letter].number_format = 'yyyy-mm-dd'
                    else:
                        worksheet.column_dimensions[col_letter].number_format = '@'
                # 设置列格式为文本，但商品金额列保持数字格式
                workbook = writer.book
                worksheet = writer.sheets['sheet1']
                for idx, col in enumerate(final_df.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    if col == '商品金额':
                        worksheet.column_dimensions[col_letter].number_format = '0.00'
                    elif col == '发货日期':
                        worksheet.column_dimensions[col_letter].number_format = 'yyyy-mm-dd'
                    else:
                        worksheet.column_dimensions[col_letter].number_format = '@'
                for col in range(1, len(final_df.columns) + 1):
                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].number_format = '@'
        except Exception as e:
            print(f"保存文件时出错: {e}")
            return
    
        
        # 处理后台sheet页
        df_backend = pd.read_excel(file_path, sheet_name='后台')
        
        # 处理实际发货单号
        if '实际发货单号' in df_backend.columns:
            df_backend['实际发货单号'] = df_backend['实际发货单号'].astype(str)
            df_backend.insert(
                df_backend.columns.get_loc('实际发货单号') + 1,
                '发货单号',
                df_backend['实际发货单号'].apply(lambda x: x[:23] if x and x != 'nan' else '')
            )
        
        # 处理发货时间和发货日期
        if '发货时间' in df_backend.columns:
            df_backend['发货日期'] = df_backend['发货时间'].astype(str).str[:10]  # 只取前10个字符，即日期部分
        
        # 保存处理后的数据
        try:
            with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                df_backend.to_excel(writer, sheet_name='后台', index=False)
        except PermissionError:
            print(f"错误: 文件 '{output_file}' 正在被其他程序使用。请关闭该文件后重试。")
            return
        except FileNotFoundError:
            try:
                df_backend.to_excel(output_file, sheet_name='后台', index=False)
            except Exception as e:
                print(f"保存文件时出错: {e}")
                return
        except Exception as e:
            print(f"处理后台数据时出错: {e}")
            return
        
        print("后台sheet页处理完成！")

    except Exception as e:
        print(f"处理后台数据时出错: {e}")
        return
# 格式化订单号和平台单号
def format_order_number(x):
    try:
        if pd.isna(x):
            return ''
        if isinstance(x, (int, float)) or (isinstance(x, str) and str(x).replace('.', '').isdigit()):
            return '{:.0f}'.format(float(x))
        return str(x)
    except:
        return str(x)

def update_order_status():
    try:
        # 第一阶段：读取数据
     
        df_original = pd.read_excel(file_path, sheet_name='sheet1')
        df_sheet1 = pd.read_excel(output_file, sheet_name='sheet1')
        df_backend = pd.read_excel(output_file, sheet_name='后台')
        
        # 转换订单号和平台单号为标准文本格式
        df_backend['订单号'] = df_backend['订单号'].apply(format_order_number)
        df_sheet1['平台单号'] = df_sheet1['平台单号'].apply(format_order_number)
        
        # 确保系统单号和运单号为文本格式
        df_sheet1['系统单号'] = df_sheet1['系统单号'].astype(str).apply(lambda x: x if x != 'nan' else '')
        # 处理运单号：转为文本并移除小数点
        df_sheet1['运单号'] = df_sheet1['运单号'].astype(str).apply(lambda x: x.replace('.', '') if x != 'nan' else '')
        
        # 创建发货仓库映射
        warehouse_map = {}
        for idx, row in df_original.iterrows():
            if isinstance(row.get('发货仓库'), str) and row['发货仓库'].startswith('4PX-'):
                warehouse_map[format_order_number(row['平台单号'])] = row['发货仓库']
        
        # 后续的映射和计算操作
        status_map = dict(zip(df_backend['订单号'], df_backend['订单状态']))
        amount_map = {}
        shipping_map = {}
        date_map = dict(zip(df_backend['订单号'], df_backend['发货日期']))
        
        # 创建订单号和订单状态的映射
        df_backend['订单号'] = df_backend['订单号'].apply(lambda x: '{:.0f}'.format(float(x)) if pd.notnull(x) and str(x).replace('.', '').isdigit() else str(x))
        status_map = dict(zip(df_backend['订单号'], df_backend['订单状态']))
        # 创建订单号和金额的映射
        amount_map = {}
        shipping_map = {}
        # 创建订单号和发货日期的映射
        date_map = dict(zip(df_backend['订单号'], df_backend['发货日期']))
        
        # 创建订单号和业务模式的映射
        remark_map = {}
        for idx, row in df_backend.iterrows():
            if row.get('订单业务模式') == '半托管仓发订单':
                remark_map[row['订单号']] = '半托管仓发订单'
            elif (row.get('订单业务模式') == '非半托管订单' and 
                  row.get('发货单号') == 'CAINIAO_STANDARD_WEIHAI'):
                remark_map[row['订单号']] = '威海仓'
        
        # 先计算金额和运费
        for idx, row in df_backend.iterrows():
            order_id = row['订单号']
            try:
                order_amount = float(row['订单金额']) if pd.notnull(row['订单金额']) else 0
                shipping_fee = float(row['物流费用']) if pd.notnull(row['物流费用']) else 0
                amount_map[order_id] = '{:.2f}'.format(order_amount - shipping_fee)  # 修改为保留2位小数
                shipping_map[order_id] = '{:.2f}'.format(shipping_fee)  # 修改为保留2位小数
            except:
                amount_map[order_id] = '0.00'  # 修改为标准格式
                shipping_map[order_id] = '0.00'  # 修改为标准格式
        
        # 然后更新sheet1中的订单状态、金额和运费
        df_sheet1['订单状态'] = df_sheet1['平台单号'].map(status_map)
        
        # 更新发货日期（仅更新空值）
        mask = df_sheet1['发货日期'].isna() | (df_sheet1['发货日期'] == '')
        df_sheet1.loc[mask, '发货日期'] = df_sheet1.loc[mask, '平台单号'].map(date_map)
        
        # 新增：如果发货日期不为空，则将状态改为"仓库已出库"
        mask_shipped = (df_sheet1['发货日期'].notna()) & (df_sheet1['发货日期'] != '')
        df_sheet1.loc[mask_shipped, '状态'] = '仓库已出库'
        
        # 更新备注（包括小包仓发货标记）
        df_sheet1['备注'] = df_sheet1['平台单号'].apply(lambda x: '小包仓发货' if str(x).endswith('-1') else '').fillna('')
        # 更新其他备注信息
        other_remarks = df_sheet1['平台单号'].str.replace('-1', '').map(remark_map)
        # 合并备注信息，保留非空的备注
        df_sheet1['备注'] = df_sheet1['备注'].combine_first(other_remarks)
        
        # 添加香港仓发货备注
        mask_hk = ((df_sheet1['发货日期'].notna()) & 
                   (df_sheet1['发货日期'] != '') & 
                   (df_sheet1['状态'] == '不发货') & 
                   ((df_sheet1['备注'].isna()) | (df_sheet1['备注'] == '')))
        df_sheet1.loc[mask_hk, '备注'] = '香港仓发货'
        
        # 更新4PX发货仓库信息到备注
        mask_4px = ((df_sheet1['备注'].isna()) | (df_sheet1['备注'] == ''))
        df_sheet1.loc[mask_4px, '备注'] = df_sheet1.loc[mask_4px, '平台单号'].map(warehouse_map)
        
        # 保存更新后的结果
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df_sheet1.to_excel(writer, sheet_name='sheet1', index=False)
                worksheet = writer.sheets['sheet1']
                if '发货日期' in df_sheet1.columns:
                    df_sheet1['发货日期'] = pd.to_datetime(df_sheet1['发货日期'], errors='coerce').dt.strftime('%Y-%m-%d')
                for idx, col in enumerate(df_sheet1.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    if col == '发货日期':
                        worksheet.column_dimensions[col_letter].number_format = 'yyyy-mm-dd'
                    else:
                        worksheet.column_dimensions[col_letter].number_format = '@'
                df_backend.to_excel(writer, sheet_name='后台', index=False)
        except PermissionError:
            print(f"错误: 文件 '{output_file}' 正在被其他程序使用。请关闭该文件后重试。")
            return
        except Exception as e:
            print(f"保存文件时出错: {e}")
            return
        
        # 重新排列列顺序，将发货日期放在第一列，并删除金额、运费、发货时间和ASIN/商品Id列
        columns = df_sheet1.columns.tolist()

        if 'ASIN/商品Id' in columns:
            columns.remove('ASIN/商品Id')
        
        df_sheet1 = df_sheet1[columns]
        
        # 最终确认系统单号和运单号为文本格式
        df_sheet1['系统单号'] = df_sheet1['系统单号'].astype(str)
        df_sheet1['运单号'] = df_sheet1['运单号'].astype(str)
        
        # 保存更新后的结果
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df_sheet1.to_excel(writer, sheet_name='sheet1', index=False)
                worksheet = writer.sheets['sheet1']
                if '发货日期' in df_sheet1.columns:
                    df_sheet1['发货日期'] = pd.to_datetime(df_sheet1['发货日期'], errors='coerce').dt.strftime('%Y-%m-%d')
                for idx, col in enumerate(df_sheet1.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    if col == '发货日期':
                        worksheet.column_dimensions[col_letter].number_format = 'yyyy-mm-dd'
                    else:
                        worksheet.column_dimensions[col_letter].number_format = '@'
                df_backend.to_excel(writer, sheet_name='后台', index=False)
                
                # 设置列格式为文本
                workbook = writer.book
                worksheet = writer.sheets['sheet1']
                
                # 找到系统单号和运单号的列索引
                col_indices = {col: idx+1 for idx, col in enumerate(df_sheet1.columns)}
                
                # 设置这些列为文本格式
                if '系统单号' in col_indices:
                    col_letter = openpyxl.utils.get_column_letter(col_indices['系统单号'])
                    worksheet.column_dimensions[col_letter].number_format = '@'
                
                if '运单号' in col_indices:
                    col_letter = openpyxl.utils.get_column_letter(col_indices['运单号'])
                    worksheet.column_dimensions[col_letter].number_format = '@'
                
        except PermissionError:
            print(f"错误: 文件 '{output_file}' 正在被其他程序使用。请关闭该文件后重试。")
            return
        except Exception as e:
            print(f"保存文件时出错: {e}")
            return
        
        print("订单状态和发货日期更新完成！")
        
    except PermissionError:
        print(f"错误: 文件正在被其他程序使用。请关闭该文件后重试。")
        return
    except Exception as e:
        print(f"更新订单状态时出错: {str(e)}")
        return

def process_warehouse_sheets():

    
    try:
        try:
            # 修改读取的 sheet 名称
            df = pd.read_excel(file_path, sheet_name='香港仓和威海仓发货', dtype=str)
            warehouse_status = [
                "交航成功", "已签收", "目的国清关完成", 
                "分拨中心出库", "已发货", "仓配交接成功"
            ]
            numeric_columns = ['用户订单号', '店铺编码', '货品ID', '商品编码']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = df[col].apply(lambda x: f'{float(x):.0f}' if pd.notnull(x) and str(x).strip() != '' else '')
            
            if '订单状态' in df.columns:
                df['订单状态'] = df['订单状态'].astype(str).apply(lambda x: ''.join(x.split()))
                clean_status = [status.replace(' ', '') for status in warehouse_status]
                mask = df['订单状态'].apply(lambda x: any(status in x for status in clean_status))
                filtered_df = df[mask]
                
                if len(filtered_df) > 0:
                    with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                        # 修改写入的 sheet 名称
                        filtered_df.rename(columns={'仓库名称': '仓名称'}, inplace=True)
                        filtered_df.to_excel(writer, sheet_name='香港仓和威海仓发货', index=False)
                        worksheet = writer.sheets['香港仓和威海仓发货']
                        for idx, col in enumerate(filtered_df.columns, 1):
                            col_letter = openpyxl.utils.get_column_letter(idx)
                            worksheet.column_dimensions[col_letter].number_format = '@'
                    
                    try:
                        df_sheet1 = pd.read_excel(output_file, sheet_name='sheet1', dtype=str)
                        mask_empty = (df_sheet1['发货仓库'].isna()) | (df_sheet1['发货仓库'] == '')
                        warehouse_mapping = dict(zip(filtered_df['用户订单号'], filtered_df['仓名称']))
                        
                        df_sheet1.loc[mask_empty, '发货仓库'] = (
                            df_sheet1.loc[mask_empty, '平台单号']
                            .map(warehouse_mapping)
                            .fillna('#N/A')
                        )
                        
                        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                            df_sheet1.to_excel(writer, sheet_name='sheet1', index=False)
                            worksheet = writer.sheets['sheet1']
                            if '发货日期' in df_sheet1.columns:
                                df_sheet1['发货日期'] = pd.to_datetime(df_sheet1['发货日期'], errors='coerce').dt.strftime('%Y-%m-%d')
                            for idx, col in enumerate(df_sheet1.columns, 1):
                                col_letter = openpyxl.utils.get_column_letter(idx)
                                if col == '发货日期':
                                    worksheet.column_dimensions[col_letter].number_format = 'yyyy-mm-dd'
                                else:
                                    worksheet.column_dimensions[col_letter].number_format = '@'
                                
                    except Exception as e:
                        raise Exception(f"更新sheet1发货仓库信息时出错 (行号: {e.__traceback__.tb_lineno}): {str(e)}")
                        
        except Exception as e:
            raise Exception(f"处理香港仓和威海仓发货sheet页时出错 (行号: {e.__traceback__.tb_lineno}): {str(e)}")
                
    except Exception as e:
        raise Exception(f"处理仓库数据时出错 (行号: {e.__traceback__.tb_lineno}): {str(e)}")
def process_semi_warehouse():

    try:
        # 处理合并单元格
        wb = openpyxl.load_workbook(file_path)
        ws = wb['半托仓发货']
        merged_ranges = list(ws.merged_cells.ranges)
        
        for merged_range in merged_ranges:
            right_value = ws.cell(merged_range.min_row, merged_range.max_col).value
            ws.unmerge_cells(str(merged_range))
            
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                ws.cell(row=row, column=merged_range.min_col, value='')
                ws.cell(row=row, column=merged_range.max_col, value=right_value)
        
        wb.save(temp_file)
        
        # 处理数据
        df = pd.read_excel(temp_file, sheet_name='半托仓发货', dtype=str)
        
        if 'Unnamed: 11' in df.columns:
            df = df.rename(columns={'Unnamed: 11': '业务类型'})
        
        unnamed_cols = [col for col in df.columns if 'Unnamed' in col]
        if unnamed_cols:
            df = df.drop(columns=unnamed_cols)
        
        df['业务类型'] = df['业务类型'].replace('TOC销售', '销售出库')
        df_filtered = df[df['业务类型'].str.contains('销售出库', case=False, na=False)]
        
        if len(df_filtered) > 0:
            with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                df_filtered.rename(columns={'仓库名称': '仓名称'}, inplace=True)
                df_filtered.to_excel(writer, sheet_name='半托仓发货', index=False)
                worksheet = writer.sheets['半托仓发货']
                for idx, col in enumerate(df_filtered.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    worksheet.column_dimensions[col_letter].number_format = '@'
            
            try:
                df_sheet1 = pd.read_excel(output_file, sheet_name='sheet1', dtype=str)
                mask_empty = (df_sheet1['发货仓库'].isna()) | (df_sheet1['发货仓库'] == '')
                warehouse_mapping = dict(zip(df_filtered['交易主单'], df_filtered['仓名称']))
                
                df_sheet1.loc[mask_empty, '发货仓库'] = (
                    df_sheet1.loc[mask_empty, '平台单号']
                    .map(warehouse_mapping)
                    .fillna('#N/A')
                )
                
                with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                    df_sheet1.to_excel(writer, sheet_name='sheet1', index=False)
                    worksheet = writer.sheets['sheet1']
                    if '发货日期' in df_sheet1.columns:
                        df_sheet1['发货日期'] = pd.to_datetime(df_sheet1['发货日期'], errors='coerce').dt.strftime('%Y-%m-%d')
                    for idx, col in enumerate(df_sheet1.columns, 1):
                        col_letter = openpyxl.utils.get_column_letter(idx)
                        if col == '发货日期':
                            worksheet.column_dimensions[col_letter].number_format = 'yyyy-mm-dd'
                        else:
                            worksheet.column_dimensions[col_letter].number_format = '@'
                        
            except Exception as e:
                raise Exception(f"更新sheet1发货仓库信息时出错 (行号: {e.__traceback__.tb_lineno}): {str(e)}")
    
    except Exception as e:
        raise Exception(f"处理半托仓发货sheet页时出错 (行号: {e.__traceback__.tb_lineno}): {str(e)}")
    
    finally:
        if os.path.exists(temp_file):
            os.remove(temp_file)
def create_check_sheet():
    try:
  
        
        # 读取后台sheet页和sheet1数据
        df_backend = pd.read_excel(output_file, sheet_name='后台', dtype=str)
        df_sheet1 = pd.read_excel(output_file, sheet_name='sheet1', dtype=str)
        
        # 创建新的DataFrame，只包含需要的列
        df_check = pd.DataFrame()
        df_check['订单号'] = df_backend['订单号']
        
        # 处理金额列，确保保留两位小数
        def clean_and_convert_amount(x):
            if pd.notnull(x) and str(x).strip() != '':
                # 移除货币符号和逗号
                cleaned_x = str(x).replace('US $', '').replace('$', '').replace(',', '').strip()
                try:
                    return '{:.2f}'.format(float(cleaned_x))
                except ValueError:
                    return '0.00' # 如果转换失败，返回 '0.00'
            return '0.00'

        df_check['订单金额'] = df_backend['订单金额'].apply(clean_and_convert_amount)
        df_check['物流费用'] = df_backend['物流费用'].apply(clean_and_convert_amount)
        
        # 计算产品金额
        df_check['产品金额'] = df_check.apply(
            lambda row: '{:.2f}'.format(float(row['订单金额']) - float(row['物流费用'])), 
            axis=1
        )
        
        # 添加sheet1的数据列
        df_check['平台单号'] = ''  # E列
        df_check['商品金额'] = ''  # F列

        
        # 填充sheet1数据
        for idx, row in df_check.iterrows():
            sheet1_row = df_sheet1[df_sheet1['平台单号'] == row['订单号']]
            if not sheet1_row.empty:
                df_check.at[idx, '平台单号'] = sheet1_row.iloc[0]['平台单号']
                df_check.at[idx, '商品金额'] = sheet1_row.iloc[0]['商品金额']
               
        
        # 新增：判断是否需要修改
        df_check['是否需要修改'] = df_check.apply(
            lambda row: '否' if (
                float(row['产品金额'] if row['产品金额'] != '' else '0.00') == 
                float(row['商品金额'] if row['商品金额'] != '' else '0.00')
            ) else '是',
            axis=1
        )
        
        # 新增：更新sheet1中的商品金额
        need_update = df_check[
            (df_check['是否需要修改'] == '是') & 
            (df_check['平台单号'].notna()) & 
            (df_check['平台单号'] != '')
        ]
        
        if not need_update.empty:
            # 创建平台单号到产品金额的映射
            update_map = dict(zip(need_update['平台单号'], need_update['产品金额']))
            
            # 更新sheet1中的商品金额
            mask = df_sheet1['平台单号'].isin(update_map.keys())
            df_sheet1.loc[mask, '商品金额'] = df_sheet1.loc[mask, '平台单号'].map(update_map)
            
            # 保存更新后的sheet1
            with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                df_sheet1.to_excel(writer, sheet_name='sheet1', index=False)
                worksheet = writer.sheets['sheet1']
                for idx, col in enumerate(df_sheet1.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    if col == '商品金额':
                        worksheet.column_dimensions[col_letter].number_format = '0.00'
                    elif col == '发货日期':
                        worksheet.column_dimensions[col_letter].number_format = 'yyyy-mm-dd'
                    else:
                        worksheet.column_dimensions[col_letter].number_format = '@'
        
        # 保存核对sheet页
        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            df_check.to_excel(writer, sheet_name='核对', index=False)
            
            # 设置列格式
            worksheet = writer.sheets['核对']
            for idx, col in enumerate(df_check.columns, 1):
                col_letter = openpyxl.utils.get_column_letter(idx)
                if col in ['订单金额', '物流费用', '产品金额', '商品金额', '商品客付运费', '实际商品金额']:
                    worksheet.column_dimensions[col_letter].number_format = '0.00'
                else:
                    worksheet.column_dimensions[col_letter].number_format = '@'
                    
    except Exception as e:
        raise Exception(f"创建核对sheet页时出错 (行号: {e.__traceback__.tb_lineno}): {str(e)}")

# 在主程序中添加新步骤
if __name__ == "__main__":
    try:
        print("=== 开始关闭WPS进程 ===")
        close_wps_processes()
        print("=== WPS进程关闭尝试完成 ===")
        
        print("\n=== Excel文件处理工具 ===")
        print("开始处理Excel文件...")
        
        print("\n第1步：处理合并单元格")
        print("✨ 正在拆分合并单元格并填充数据...")
        unmerge_and_fill()
        print("✅ 合并单元格处理完成")
        
        print("\n第2步：处理后台数据")
        print("✨ 正在处理后台sheet页数据...")
        print("• 处理订单号格式")
        print("• 处理发货单号")
        print("• 更新发货日期")
        process_backend()
        print("✅ 后台数据处理完成")
        
        print("\n第3步：更新订单状态")
        print("✨ 正在更新订单状态和相关信息...")
        print("• 更新订单状态")
        print("• 更新发货日期")
        print("• 处理备注信息")
        print("• 更新仓库信息")
        update_order_status()
        print("✅ 订单状态更新完成")
        
        print("\n第4步：处理仓库数据")
        print("✨ 正在处理香港仓和威海仓发货数据...")
        process_warehouse_sheets()
        print("✅ 仓库数据处理完成")
        
        print("\n第5步：处理半托仓发货数据")
        print("✨ 正在处理半托仓发货sheet页...")
        process_semi_warehouse()  # 直接调用函数即可
        print("✅ 半托仓发货数据处理完成")
        
        # 新增第6步
        print("\n第6步：创建核对sheet页")
        print("✨ 正在创建核对sheet页...")
        create_check_sheet()
        print("✅ 核对sheet页创建完成")
        
        print("\n=== 处理完成 ===")
        print("✅ 所有数据处理已完成")
        print("📁 结果已保存至：速卖通订单管理202502-处理后.xlsx")
        print("\n如需查看详细处理日志，请查看上方输出信息。")
        
    except Exception as e:
        import traceback
        error_info = traceback.extract_tb(e.__traceback__)[-1]
        file_name = error_info.filename.split('\\')[-1]
        line_no = error_info.lineno
        print("\n=== 处理出错 ===")
        print(f"❌ 错误位置: {file_name}, 第 {line_no} 行")
        print(f"❌ 错误信息: {str(e)}")
        print("建议：请检查Excel文件格式是否正确，以及文件是否被其他程序占用。")
    

