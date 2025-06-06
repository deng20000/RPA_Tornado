import xbot
from xbot import print, sleep
from .import package
from .package import variables as glv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def main(args):
    use_data()

def use_data():
    try:
        # 获取输入文件路径
        # inputfile = r"C:\Users\gl-02251756\Downloads\2025-04-26每日数据Shareasale账户数据查看.csv"
        inputfile = glv["g_file"]
        print(f"Input file: {inputfile}")

        # 1. 使用 pandas 读取 CSV 文件，处理格式问题
        try:
            df = pd.read_csv(inputfile, encoding='utf-8', on_bad_lines='skip')
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(inputfile, encoding='gbk', on_bad_lines='skip')
            except UnicodeDecodeError:
                df = pd.read_csv(inputfile, encoding='iso-8859-1', on_bad_lines='skip')

        print("CSV file read successfully.")

        # 2. 保留指定的列
        columns_to_keep = [
            'transID', 'userID', 'transdate', 'transamount', 'commission', 'ssamount', 
            'voided', 'locked', 'pending', 'lastip', 'lastreferer', 'skulist', 
            'priceList', 'quantityList', 'orderNumber', 'couponCode', 'userAgent', 
            'usedACoupon'
        ]
        
        # 检查列是否存在
        missing_columns = [col for col in columns_to_keep if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Missing columns in CSV file: {missing_columns}")
        
        df = df[columns_to_keep]
        print("Columns filtered successfully.")

        # 3. 将 DataFrame 写入 Excel 文件
        output_file = inputfile.replace('.csv', '.xlsx')
        df.to_excel(output_file, index=False)
        print(f"Data written to Excel file: {output_file}")

        # 4. 使用 openpyxl 打开 Excel 文件并对符合条件的单元格标红
        wb = load_workbook(output_file)
        ws = wb.active

        # 定义标红的填充样式
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # 获取 "userAgent" 列的索引
        user_agent_col_index = columns_to_keep.index('userAgent') + 1  # Excel 列索引从 1 开始

        # 遍历 "userAgent" 列
        for row in ws.iter_rows(min_row=2, min_col=user_agent_col_index, max_col=user_agent_col_index):
            for cell in row:
                user_agent = cell.value
                if user_agent and any(keyword in user_agent.lower() for keyword in ['bot', 'spider', 'crawler', 'scraper', 'monitor', 'indexer']):
                    cell.fill = red_fill

        # 保存修改后的 Excel 文件
        wb.save(output_file)
        print("Excel file saved with highlighted cells.")

    except Exception as e:
        print(f"Error: {e}")
        raise  # 重新抛出异常以便调试

# glv["g_file"] = r"C:\Users\GL\Downloads\2025-04-26每日数据Shareasale账户数据查看.csv"
# main("")