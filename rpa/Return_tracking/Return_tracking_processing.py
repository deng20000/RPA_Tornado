"""
退货追踪数据处理模块

该模块用于处理退货订单数据，将JSON格式的订单数据按地区和平台分类，
并导出为Excel文件，支持亚马逊和Shopify两个平台的订单处理。

主要功能：
1. 解析JSON格式的订单数据
2. 按地区（美国、欧洲、英国）和平台（亚马逊、Shopify）分类订单
3. 导出分类后的数据到Excel文件

作者：[作者名称]
创建时间：[创建时间]
最后修改：[修改时间]
"""

import json
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import os


# ==================== 全局变量定义 ====================

# 初始化结果列表 - 用于存储不同地区和平台的订单数据
result_US = []          # 亚马逊美国订单列表
result_EU = []          # 亚马逊欧洲订单列表
result_UK = []          # 亚马逊英国订单列表
result_shopfiy_US = []  # Shopify美国订单列表
result_shopfiy_EU = []  # Shopify欧洲订单列表
result_shopfiy_UK = []  # Shopify英国订单列表

# 定义亚马逊欧洲国家代码列表
AMAZON_EU_COUNTRIES = ['IT', 'DE', 'FR', 'ES', 'NL', 'SE', 'TR', 'PL', 'BE']


# ==================== 工具函数 ====================

def safe_get(data, keys, default=''):
    """
    安全获取嵌套字典的值
    
    Args:
        data (dict): 要查询的字典数据
        keys (list): 键的路径列表
        default: 默认返回值
        
    Returns:
        获取到的值或默认值
        
    Example:
        safe_get({'a': {'b': 'value'}}, ['a', 'b'], '') -> 'value'
        safe_get({'a': {}}, ['a', 'b'], 'default') -> 'default'
    """
    if not isinstance(data, dict):
        return default
        
    for key in keys:
        if key in data:
            data = data[key]
        else:
            return default
    return data

# ==================== 核心处理函数 ====================

def process_orders(input_data):
    """
    处理订单数据并按地区和平台分类
    
    该函数可以接受JSON文件路径或直接传入数据列表，解析每个订单的详细信息，
    并根据店铺名称将订单分类到不同的地区和平台列表中。
    
    Args:
        input_data (str or list): 输入数据，可以是：
            - str: JSON文件路径
            - list: 订单数据列表（格式与JSON文件中的数据相同）
        
    Returns:
        dict: 包含各地区和平台订单数据的字典
            - US_orders: 亚马逊美国订单列表
            - EU_orders: 亚马逊欧洲订单列表  
            - UK_orders: 亚马逊英国订单列表
            - shopfiy_US: Shopify美国订单列表
            - shopfiy_EU: Shopify欧洲订单列表
            - shopfiy_UK: Shopify英国订单列表
            
    Raises:
        ValueError: 当文件内容为空或数据格式错误时
        json.JSONDecodeError: 当JSON格式错误时
        FileNotFoundError: 当文件不存在时
        TypeError: 当输入数据类型不支持时
    """
    
    # 判断输入数据类型并进行相应处理
    if isinstance(input_data, str):
        # 输入是文件路径，读取JSON文件
        print(f"开始处理文件: {input_data}")
        try:
            # 读取并验证JSON文件
            with open(input_data, 'r', encoding='utf-8') as f:
                # 先读取文件内容进行验证
                raw_data = f.read()
                if not raw_data.strip():
                    raise ValueError("文件内容为空")
                    
                # 替换可能的非法字符，确保JSON格式正确
                cleaned_data = raw_data.replace('\ufeff', '')  # 移除BOM头
                
                try:
                    data = json.loads(cleaned_data)
                except json.JSONDecodeError as e:
                    print(f"JSON解析错误: {str(e)}")
                    print(f"错误位置: 行{e.lineno}, 列{e.colno}")
                    print(f"附近内容: {raw_data[max(0,e.pos-20):e.pos+20]}")
                    raise
            print(f"成功加载{len(data)}条记录")
            
        except FileNotFoundError:
            print(f"文件未找到: {input_data}")
            raise
            
    elif isinstance(input_data, list):
        # 输入是数据列表，直接使用
        print(f"开始处理传入的数据列表")
        data = input_data
        print(f"成功接收{len(data)}条记录")
        
        # 验证数据格式
        if not data:
            raise ValueError("传入的数据列表为空")
            
        # 简单验证第一条记录的格式
        if data and not isinstance(data[0], dict):
            raise ValueError("数据格式错误：列表中的元素应为字典格式")
            
        if data and 'fields' not in data[0]:
            raise ValueError("数据格式错误：缺少必需的'fields'字段")
            
    else:
        # 不支持的数据类型
        raise TypeError(f"不支持的输入数据类型: {type(input_data)}。请传入文件路径(str)或数据列表(list)")
    
    # 遍历处理每个订单记录
    for idx, item in enumerate(data, 1):
        try:
            print(f"\n处理第{idx}条记录...")
            print(f"记录ID: {item.get('id', '无ID')}")
            
            # 提取订单字段信息
            fields = item.get('fields', {})
            print(f"字段数量: {len(fields)}")
            
            # 提取Shopify站点名称 - 支持多种可能的字段名称格式
            shopify_site = ''
            site_field = None
            
            # 尝试不同可能的字段名称（考虑大小写和命名差异）
            if 'shopify站点' in fields:
                site_field = 'shopify站点'
                shopify_site = fields['shopify站点'][0].get('name', '') if isinstance(fields['shopify站点'], list) and fields['shopify站点'] else ''
            elif 'Shopify站点' in fields:
                site_field = 'Shopify站点'
                shopify_site = fields['Shopify站点'][0].get('name', '') if isinstance(fields['Shopify站点'], list) and fields['Shopify站点'] else ''
            elif 'shopify_site' in fields:
                site_field = 'shopify_site'
                shopify_site = fields['shopify_site'][0].get('name', '') if isinstance(fields['shopify_site'], list) and fields['shopify_site'] else ''
            
            print(f"使用的站点字段: {site_field}, 站点值: {shopify_site}")
            
            # 检查站点信息是否存在，跳过无站点信息的订单
            if not shopify_site:
                print(f"跳过无站点信息的订单")
                continue
            
            # 构建订单数据数组 - 按照Excel表格的列顺序排列
            order_data = [
                '退货统计表',  # 表中来源数据（固定值）
                shopify_site if shopify_site else '',  # 店铺名称
                fields.get('订单号', ''),  # 订单号
                # 产品型号 - 处理可能的列表格式
                fields.get('产品型号', [{}])[0].get('name', '') if isinstance(fields.get('产品型号'), list) else '',
                fields.get('数量', ''),  # 订单数量
                # 购买渠道 - 提取name字段
                fields.get('购买渠道', {}).get('name', ''),
                # 退货时间 - 从时间戳转换为日期格式
                datetime.fromtimestamp(int(fields.get('退货时间', '0'))/1000).strftime('%Y-%m-%d') if fields.get('退货时间') else '',
                # 购买时间 - 从时间戳转换为日期格式  
                datetime.fromtimestamp(int(fields.get('购买时间', '0'))/1000).strftime('%Y-%m-%d') if fields.get('购买时间') else '',
                fields.get('退货点', ''),  # 退货点
                fields.get('客户是否寄回(寄回单号）', '')  # 客户是否寄回商品
            ]
            
            # 根据店铺名称进行地区和平台分类
            # 亚马逊平台分类
            if 'Amazon UK' in shopify_site:
                result_UK.append(order_data)
            elif 'Amazon US' in shopify_site:
                result_US.append(order_data)
            elif any(f'Amazon {country}' in shopify_site for country in AMAZON_EU_COUNTRIES):
                result_EU.append(order_data)
            # Shopify平台分类
            elif 'UK' in shopify_site and 'Amazon' not in shopify_site:
                result_shopfiy_UK.append(order_data)
            elif 'US' in shopify_site and 'Amazon' not in shopify_site:
                result_shopfiy_US.append(order_data)
            elif any(f'{country}' in shopify_site for country in AMAZON_EU_COUNTRIES) and 'Amazon' not in shopify_site:
                result_shopfiy_EU.append(order_data)
            else:
                print(f"跳过非目标订单: {shopify_site}")
                
        except Exception as e:
            print(f"处理订单时出错: {e}")
            continue
    
    return {
        'US_orders': result_US,
        'EU_orders': result_EU,
        'UK_orders': result_UK,
        'shopfiy_US': result_shopfiy_US,
        'shopfiy_EU': result_shopfiy_EU,
        'shopfiy_UK': result_shopfiy_UK
    }


# ==================== Excel导出函数 ====================

def export_to_excel(results):
    """
    将分类后的订单数据导出到Excel文件
    
    该函数为每个地区（美国、欧洲、英国）创建独立的Excel工作簿，
    每个工作簿包含三个工作表：亚马逊买家退货、Shopify买家退货、亚马逊移除
    
    Args:
        results (dict): 包含各地区和平台订单数据的字典
            - US_orders: 亚马逊美国订单列表
            - EU_orders: 亚马逊欧洲订单列表
            - UK_orders: 亚马逊英国订单列表
            - shopfiy_US: Shopify美国订单列表
            - shopfiy_EU: Shopify欧洲订单列表
            - shopfiy_UK: Shopify英国订单列表
            
    Returns:
        None
        
    生成文件格式：
        - 文件名：YYYY年MM月DD日退货追踪(地区名).xlsx
        - 工作表：
          1. 亚马逊-买家退货-地区名
          2. shopfiy-地区名买家退货  
          3. 亚马逊移除-地区名（空白表）
    """
    try:
        # 定义Excel表格的列标题
        headers = [
            '表中来源数据', '店铺', '订单号', '产品型号', '数量',
            '购买渠道', '退货时间', '购买时间', '退货点', '客户是否寄回'
        ]
        
        # 生成当前日期作为文件名的一部分
        today = datetime.now().strftime('%Y年%m月%d日')
        current_dir = os.path.dirname(os.path.abspath(__file__))
        
        # 定义地区映射关系（英文代码 -> 中文名称）
        regions = {
            'US': '美国',
            'EU': '欧洲', 
            'UK': '英国'
        }
        
        # 为每个地区创建独立的Excel工作簿
        for region_code, region_name in regions.items():
            print(f"\n正在处理{region_name}地区数据...")
            print(f"亚马逊订单数量: {len(results[f'{region_code}_orders'])}")
            print(f"Shopify订单数量: {len(results[f'shopfiy_{region_code}'])}")
            
            # 创建新的Excel工作簿
            wb = Workbook()
            print(f"初始sheet列表: {wb.sheetnames}")
            
            # 删除默认创建的空白工作表
            if wb.sheetnames:
                default_sheet = wb.active
                print(f"删除默认sheet: {default_sheet.title}") # type: ignore
                wb.remove(default_sheet) # type: ignore
            
            print(f"处理后sheet列表: {wb.sheetnames}")
            
            # 创建亚马逊买家退货工作表
            amazon_sheet = wb.create_sheet(f"亚马逊-买家退货-{region_name}")
            amazon_sheet.append(headers)  # 添加表头
            
            # 填充亚马逊订单数据
            amazon_count = 0
            for row in results[f'{region_code}_orders']:
                amazon_sheet.append(row)
                amazon_count += 1
            print(f"已写入{amazon_count}条亚马逊订单")
            
            # 创建Shopify买家退货工作表
            shopify_sheet = wb.create_sheet(title=f"shopfiy-{region_name}买家退货")
            shopify_sheet.append(headers)  # 添加表头
            
            # 填充Shopify订单数据
            shopify_count = 0
            for row in results[f'shopfiy_{region_code}']:
                shopify_sheet.append(row)
                shopify_count += 1
            print(f"已写入{shopify_count}条Shopify订单")
            
            # 创建空白的亚马逊移除工作表（预留用于手动填写）
            removal_sheet = wb.create_sheet(title=f"亚马逊移除-{region_name}")
            removal_sheet.append(headers)  # 只添加表头，数据部分留空
            print(f"已创建sheet: {removal_sheet.title}")
            
            # 保存文件前验证所有工作表
            print("\n最终sheet列表:")
            for sheet in wb.sheetnames:
                print(f"- {sheet}")
                sheet_obj = wb[sheet]
                print(f"  行数: {sheet_obj.max_row}, 列数: {sheet_obj.max_column}")
            
            # 生成文件名并保存Excel文件
            filename = os.path.join(current_dir, f"{today}退货追踪({region_name}).xlsx")
            wb.save(filename)
            print(f"\n已成功生成文件: {filename}")
            print(f"文件路径: {os.path.abspath(filename)}")
            print("请检查该路径下的Excel文件")
            
    except Exception as e:
        print(f"导出Excel时发生错误: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        print("导出Excel完成")

# ==================== 主程序入口 ====================

if __name__ == "__main__":
    """
    主程序入口
    
    执行流程：
    1. 处理JSON订单数据文件或数据列表
    2. 按地区和平台分类订单
    3. 输出统计信息和示例数据
    4. 导出分类结果到Excel文件
    """
    
    # 方式1：处理订单数据文件（默认文件名：test.json）
    print("=== 开始处理退货订单数据（从文件） ===")
    results = process_orders('test.json')
    
    # 方式2：直接传入数据列表（示例）
    # sample_data = [
    #     {
    #         "订单号": "12345",
    #         "产品型号": [{"name": "产品A"}],
    #         "数量": 1,
    #         "购买渠道": {"name": "Amazon UK"},
    #         "退货时间": "1640995200000",
    #         "购买时间": "1640908800000",
    #         "退货点": "UK仓库",
    #         "客户是否寄回(寄回单号）": "已寄回"
    #     }
    # ]
    # print("=== 开始处理退货订单数据（从列表） ===")
    # results = process_orders(sample_data)
    
    # 输出初步统计结果
    print(f"亚马逊美国订单数量: {len(results['US_orders'])}")
    print(f"亚马逊欧洲订单数量: {len(results['EU_orders'])}")
    print(f"亚马逊英国订单数量: {len(results['UK_orders'])}")
    print(f"Shopify美国订单数量: {len(results['shopfiy_US'])}")
    print(f"Shopify欧洲订单数量: {len(results['shopfiy_EU'])}")
    print(f"Shopify英国订单数量: {len(results['shopfiy_UK'])}")
    
    # 导出分类结果到Excel文件
    print("\n=== 开始导出Excel文件 ===")
    export_to_excel(results)
    
    # 输出详细统计结果
    print("\n=== 订单统计结果 ===")
    print(f"亚马逊美国订单数量: {len(results['US_orders'])}")
    print(f"亚马逊欧洲订单数量: {len(results['EU_orders'])}")
    print(f"亚马逊英国订单数量: {len(results['UK_orders'])}")
    print(f"Shopify美国订单数量: {len(results['shopfiy_US'])}")
    print(f"Shopify欧洲订单数量: {len(results['shopfiy_EU'])}")
    print(f"Shopify英国订单数量: {len(results['shopfiy_UK'])}")

    # 定义表头用于示例数据展示
    headers = [
        '表中来源数据', '店铺', '订单号', '产品型号', '数量',
        '购买渠道', '退货时间', '购买时间', '退货点', '客户是否寄回'
    ]
    
    # 输出各类订单的示例数据（用于验证数据格式）
    print("\n=== 订单示例数据 ===")
    
    print("美国订单示例(带字段名):")
    if results['US_orders']:
        print(headers)
        print(results['US_orders'][0])
    else:
        print("无美国订单数据")
    
    print("\n欧洲订单示例(带字段名):")
    if results['EU_orders']:
        print(headers)
        print(results['EU_orders'][0])
    else:
        print("无欧洲订单数据")
    
    print("\n亚马逊英国订单示例(带字段名):")
    if results['UK_orders']:
        print(headers)
        print(results['UK_orders'][0])
    else:
        print("无英国订单数据")
    
    print("\nShopify美国订单示例(带字段名):")
    if results['shopfiy_US']:
        print(headers)
        print(results['shopfiy_US'][0])
    else:
        print("无Shopify美国订单数据")
    
    print("\nShopify欧洲订单示例(带字段名):")
    if results['shopfiy_EU']:
        print(headers)
        print(results['shopfiy_EU'][0])
    else:
        print("无Shopify欧洲订单数据")
    
    print("\nShopify英国订单示例(带字段名):")
    if results['shopfiy_UK']:
        print(headers)
        print(results['shopfiy_UK'][0])
    else:
        print("无Shopify英国订单数据")
    
    print("\n=== 处理完成 ===")
    print("请查看生成的Excel文件以获取完整的分类结果")
        

    