import pandas as pd
import openpyxl
import re
from datetime import datetime



def process_google_ads_data(file_path):
    result_dict = {}
    result_dict["地区"] = []
    try:
        # 使用openpyxl查找包含"GL-iNet"的sheet
        # wb = openpyxl.load_workbook(file_path)
        # 使用pandas读取excel
        sheets = get_all_sheets(file_path)
        for sheet in sheets:
            df = pd.read_excel(file_path, sheet_name=sheet)
            # 查找 第一列 最后一个 包含 总计：广告系列 的行,默认从第4行开始
            last_row = df[df.iloc[:, 0].str.contains("地区")].index[-1]
            print(last_row)
            # 当天数据
            data = df.iloc[3:last_row]
            # 地区数据
            region_data = df.iloc[last_row+1:]
            # 在最左侧插入一列日期
            data.insert(0, '日期', datetime.now().strftime('%Y-%m-%d'))
            region_data.insert(0, '日期', datetime.now().strftime('%Y-%m-%d'))
            # 把nan替换为空字符串
            data = data.fillna("")
            region_data = region_data.fillna("")
            
            # 当两个数据转为二维列表
            data = data.values.tolist()
            region_data = region_data.values.tolist()
            base_name = get_base_sheet_name(sheet)
            result_dict[base_name] = data
            result_dict[base_name+"地区"] = region_data
            result_dict["地区"] = result_dict["地区"]+ region_data
            # print(data)
            # print(region_data)
        print(result_dict["地区"])
        return result_dict
            
    except Exception as e:
        print(f"处理Excel文件时发生错误: {e}")
        import traceback
        traceback.print_exc()
        return [], []

# Helper function to get base sheet name
def get_base_sheet_name(sheet_name):
    # Remove date patterns like -YYYY-MM-DD or YYYY年MM月DD日
    base_name = re.sub(r'[-_]?\d{4}[-.]\d{2}[-.]\d{2}$', '', sheet_name)
    base_name = re.sub(r'\d{4}年\d{2}月\d{2}日$', '', base_name)
    return base_name.strip()

print(get_base_sheet_name("GL-iNet-US-2025-06-11"))

# 获取到所有sheet页
def get_all_sheets(file_path):
    wb = openpyxl.load_workbook(file_path)
    filiter = [sheet for sheet in wb.sheetnames if "GL-iNet" in sheet]
    # 过滤掉不带当天日期的
    return filiter

print(get_all_sheets("c:\\Users\\gl-02251756\\Desktop\\rpa\\google_watch\\2025-06-11 Google Ads数据查看.xlsx"))

if __name__ == "__main__":
    excel_file = "c:\\Users\\gl-02251756\\Desktop\\rpa\\google_watch\\2025-06-11 Google Ads数据查看.xlsx"
    result_dict = process_google_ads_data(excel_file)

    print(result_dict.keys())
    # print(result_dict["GL-iNet-US"])
    # print(result_dict["GL-iNet-US地区"])
    
        