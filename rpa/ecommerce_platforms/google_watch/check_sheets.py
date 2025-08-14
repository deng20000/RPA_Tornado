import openpyxl

def check_excel_sheets(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        print('Excel文件中的sheet名称:')
        for sheet in wb.sheetnames:
            print(f'- {sheet}')
            
        # 检查第一个sheet的前几行内容
        first_sheet = wb[wb.sheetnames[0]]
        print(f'\n第一个sheet "{wb.sheetnames[0]}" 的前5行内容:')
        for r in range(1, min(6, first_sheet.max_row + 1)):
            row_values = []
            for c in range(1, min(5, first_sheet.max_column + 1)):
                cell_value = first_sheet.cell(row=r, column=c).value
                row_values.append(str(cell_value))
            print(f'行 {r}: {" | ".join(row_values)}')
            
    except Exception as e:
        print(f'检查Excel文件时出错: {str(e)}')

if __name__ == '__main__':
    excel_file = "c:\\Users\\gl-02251756\\Desktop\\rpa\\google_watch\\2025-06-11 Google Ads数据查看.xlsx"
    check_excel_sheets(excel_file)