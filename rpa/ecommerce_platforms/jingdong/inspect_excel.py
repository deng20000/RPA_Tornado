import openpyxl

def inspect_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        print(f'Sheet names: {wb.sheetnames}')
        
        sheet = wb.active
        print(f'Active sheet: {sheet.title}')
        
        print('\nFirst 10 rows sample:')
        for r in range(1, min(11, sheet.max_row + 1)):
            print(f'\nRow {r}:')
            for c in range(1, min(10, sheet.max_column + 1)):
                cell_value = sheet.cell(row=r, column=c).value
                print(f'  Col {c}: {cell_value}')
                
        # 特别检查包含"订单编号:"的单元格
        print('\n\n查找包含"订单编号:"的单元格:')
        found = False
        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=r, column=c).value
                if cell_value and '订单编号:' in str(cell_value):
                    print(f'找到匹配单元格 - 行: {r}, 列: {c}, 值: {cell_value}')
                    found = True
        
        if not found:
            print('未找到包含"订单编号:"的单元格')
            
    except Exception as e:
        print(f'Error: {str(e)}')

if __name__ == '__main__':
    excel_file = "c:\\Users\\gl-02251756\\Desktop\\rpa\\jingdong\\2025年06月10日京东数据明细.xlsx"
    inspect_excel(excel_file)