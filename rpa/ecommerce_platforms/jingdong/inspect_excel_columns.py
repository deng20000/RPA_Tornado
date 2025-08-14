import openpyxl

def inspect_excel_columns(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        print(f'Sheet names: {wb.sheetnames}')
        
        sheet = wb.active
        print(f'Active sheet: {sheet.title}')
        
        print('\n检查Q-W列的内容:')
        for r in range(1, min(20, sheet.max_row + 1)):
            row_has_data = False
            row_data = []
            
            for c in range(17, 24):  # Q=17, W=23
                cell_value = sheet.cell(row=r, column=c).value
                if cell_value:
                    row_has_data = True
                row_data.append(f'Col {c} ({chr(64+c)}): {cell_value}')
            
            if row_has_data:
                print(f'\nRow {r}:')
                for cell_info in row_data:
                    print(f'  {cell_info}')
        
    except Exception as e:
        print(f'Error: {str(e)}')

if __name__ == '__main__':
    excel_file = "c:\\Users\\gl-02251756\\Desktop\\rpa\\jingdong\\2025年06月10日京东数据明细.xlsx"
    inspect_excel_columns(excel_file)